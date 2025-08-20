import argparse
import datetime
import pandas as pd
import os

def valid_date(s):
    """Validate and parse date string in YYYY-MM-DD format."""
    try:
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        msg = f"Not a valid date: '{s}'. Expected format: YYYY-MM-DD."
        raise argparse.ArgumentTypeError(msg)

def parse_arguments():
    """Parse command-line arguments for the script."""
    parser = argparse.ArgumentParser(description="Filter and process order data from a CSV file.")
    parser.add_argument(
        '--start_date',
        required=True,
        type=valid_date,
        help="Start date for filtering orders (format: YYYY-MM-DD)."
    )
    parser.add_argument(
        '--end_date',
        required=True,
        type=valid_date,
        help="End date for filtering orders (format: YYYY-MM-DD)."
    )
    return parser.parse_args()

def main():
    """Main function to run the script."""
    args = parse_arguments()
    print(f"Script started. Filtering orders from {args.start_date} to {args.end_date}")

    input_file = 'Example CSV/orders_export_1.csv'

    # Check for input file existence
    if not os.path.exists(input_file):
        print(f"Error: Input file not found at '{input_file}'")
        return

    # Load the CSV data
    try:
        df = pd.read_csv(input_file)
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return

    print(f"Successfully loaded {len(df)} rows from {input_file}")

    # Identify and separate problematic orders
    critical_columns = ['Fulfilled at', 'Lineitem sku', 'Lineitem quantity', 'Lineitem price', 'Name']

    # Verify that critical columns exist
    missing_cols = [col for col in critical_columns if col not in df.columns]
    if missing_cols:
        print(f"Error: The following critical columns are missing from the input file: {', '.join(missing_cols)}")
        return

    # Find rows with nulls in the specified columns (excluding 'Name' for now)
    check_cols = ['Fulfilled at', 'Lineitem sku', 'Lineitem quantity', 'Lineitem price']
    rows_with_nulls = df[df[check_cols].isnull().any(axis=1)]

    problematic_order_ids = rows_with_nulls['Name'].unique()

    problematic_orders_df = df[df['Name'].isin(problematic_order_ids)].copy()
    main_df = df[~df['Name'].isin(problematic_order_ids)].copy()

    print(f"Identified {len(problematic_order_ids)} problematic orders with {len(problematic_orders_df)} total line items.")
    print(f"Processing {len(main_df)} remaining line items.")

    # Step 4: Filter the main data
    # Filter by fulfillment status
    if 'Fulfillment Status' not in main_df.columns:
        print("Error: 'Fulfillment Status' column not found.")
        return

    filtered_df = main_df[main_df['Fulfillment Status'] == 'fulfilled'].copy()
    print(f"Filtered by status 'fulfilled'. {len(filtered_df)} rows remaining for processing.")

    # Convert 'Fulfilled at' to datetime objects and handle conversion errors
    # Using utc=True ensures a consistent datetime type, avoiding errors with mixed timezones.
    filtered_df['Fulfilled at DT'] = pd.to_datetime(filtered_df['Fulfilled at'], errors='coerce', utc=True)

    # Identify orders where date conversion failed (resulted in NaT)
    failed_conversion_mask = filtered_df['Fulfilled at DT'].isnull()
    if failed_conversion_mask.any():
        failed_orders_ids = filtered_df[failed_conversion_mask]['Name'].unique()

        # Move these newly identified problematic orders to the problematic_orders_df
        newly_problematic_df = filtered_df[filtered_df['Name'].isin(failed_orders_ids)]
        # Ensure columns match before concatenating
        problematic_orders_df = pd.concat([problematic_orders_df, newly_problematic_df.drop(columns=['Fulfilled at DT'])])

        # Remove them from the main processing dataframe
        filtered_df = filtered_df.dropna(subset=['Fulfilled at DT'])
        print(f"Moved {len(failed_orders_ids)} orders with unreadable dates to the problematic list.")

    # Filter by the specified date range (inclusive)
    # The .dt.date accessor converts the datetime objects to date objects for comparison
    filtered_df['Fulfilled at Date'] = filtered_df['Fulfilled at DT'].dt.date
    date_mask = (filtered_df['Fulfilled at Date'] >= args.start_date) & (filtered_df['Fulfilled at Date'] <= args.end_date)

    final_filtered_df = filtered_df[date_mask].copy()

    # Clean up the temporary columns used for filtering
    final_filtered_df.drop(columns=['Fulfilled at DT', 'Fulfilled at Date'], inplace=True, errors='ignore')

    print(f"Filtered by date range. {len(final_filtered_df)} final rows selected.")

    # Step 5: Create and style the output XLSX file
    output_filename = f"filtered_orders_{args.start_date.strftime('%Y-%m-%d')}_{args.end_date.strftime('%Y-%m-%d')}.xlsx"

    # Define the columns for the final report, including 'Fulfillment Status' for styling
    output_columns = ['Name', 'Fulfilled at', 'Fulfillment Status', 'Lineitem sku', 'Lineitem quantity', 'Lineitem price']

    # Ensure final_filtered_df has all the necessary columns before proceeding
    final_filtered_df = final_filtered_df.reindex(columns=output_columns)

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # --- Write and Style Filtered Orders Sheet ---
            if not final_filtered_df.empty:
                final_filtered_df.to_excel(writer, sheet_name='Filtered Orders', index=False)
                worksheet = writer.sheets['Filtered Orders']

                from openpyxl.styles import Font, Border, Side
                from openpyxl.utils import get_column_letter

                bold_font = Font(bold=True)
                thin_top_border = Border(top=Side(style='thin'))

                # Get column indices for styling
                col_indices = {name: i + 1 for i, name in enumerate(final_filtered_df.columns)}
                name_col_idx = col_indices.get('Name')
                status_col_idx = col_indices.get('Fulfillment Status')

                # Style header
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=1, column=col_idx).font = bold_font

                # Apply styles row by row
                for row_idx in range(2, worksheet.max_row + 1):
                    # Apply bold font to specific columns
                    if name_col_idx:
                        worksheet.cell(row=row_idx, column=name_col_idx).font = bold_font
                    if status_col_idx:
                        worksheet.cell(row=row_idx, column=status_col_idx).font = bold_font

                    # Apply border between orders
                    if name_col_idx and row_idx > 1:
                        current_order_name = worksheet.cell(row=row_idx, column=name_col_idx).value
                        prev_order_name = worksheet.cell(row=row_idx - 1, column=name_col_idx).value
                        if current_order_name != prev_order_name:
                            for col_idx in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_idx, column=col_idx).border = thin_top_border

                # Auto-fit columns
                for col_idx, column in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # --- Write and Style Problematic Orders Sheet ---
            if not problematic_orders_df.empty:
                problematic_orders_df.to_excel(writer, sheet_name='Problematic Orders', index=False)
                worksheet_prob = writer.sheets['Problematic Orders']
                # Auto-fit columns
                for col_idx, column in enumerate(worksheet_prob.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet_prob.column_dimensions[column_letter].width = adjusted_width

        print(f"Processing complete. Output saved to '{output_filename}'")

    except Exception as e:
        print(f"An error occurred while writing the Excel file: {e}")

if __name__ == "__main__":
    main()
